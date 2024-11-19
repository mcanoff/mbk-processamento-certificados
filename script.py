import os
import re
import json
import shutil
import string
import logging
import pandas as pd
from requests import request
from winotify import Notification
from datetime import datetime, timedelta
from openpyxl import load_workbook
from mbk_email import Email, EmailTest, process_file_info
from digital_certificate.cert import Certificate

# Diretórios para salvar os arquivos
XLSX_OUTPUT_DIR = r"C:\Users\miria\OneDrive\Área de Trabalho\mbk\xlsx"
LOG_DIR = r"C:\Users\miria\OneDrive\Área de Trabalho\mbk\logs"
VENCIDOS_DIR = r"C:\Users\miria\OneDrive\Área de Trabalho\mbk\vencidos"

# Configuração de logs
if not os.path.exists(LOG_DIR):
    os.makedirs(LOG_DIR)
log_filename = os.path.join(LOG_DIR, f"processamento_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log")
logging.basicConfig(filename=log_filename, level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Variáveis para popular tabela
ROOT_DIR = r"C:\Users\miria\OneDrive\Área de Trabalho\mbk"
PWD_DIR = r"C:\Users\miria\OneDrive\Área de Trabalho\mbk\pwd"

# Carregar planilhas existentes
cnpj_wb = load_workbook(os.path.join(ROOT_DIR, "Clientes ativos - PJ.xlsx"))
cnpj_ws = cnpj_wb.active

cpf_wb = load_workbook(os.path.join(ROOT_DIR, "Clientes ativos - PF.xlsx"))
cpf_ws = cpf_wb.active

inativos_wb = load_workbook(os.path.join(ROOT_DIR, "Clientes inativos - PJ e PF.xlsx"))
inativos_ws = inativos_wb.active

# Variáveis para enviar email e notificação
today = datetime.today()
day_15 = today + timedelta(days=15)
day_14 = today + timedelta(days=14)
day_13 = today + timedelta(days=13)
day_12 = today + timedelta(days=12)
day_11 = today + timedelta(days=11)
day_10 = today + timedelta(days=10)
day_9 = today + timedelta(days=9)
day_8 = today + timedelta(days=8)
day_7 = today + timedelta(days=7)
day_6 = today + timedelta(days=6)
day_5 = today + timedelta(days=5)
day_4 = today + timedelta(days=4)
day_3 = today + timedelta(days=3)
day_2 = today + timedelta(days=2)
day_1 = today + timedelta(days=1)

counter = {
    "15": day_15,
    "14": day_14,
    "13": day_13,
    "12": day_12,
    "11": day_11,
    "10": day_10,
    "9": day_9,
    "8": day_8,
    "7": day_7,
    "6": day_6,
    "5": day_5,
    "4": day_4,
    "3": day_3,
    "2": day_2,
    "1": day_1
}

PWD = open(os.path.join(PWD_DIR, "pwd.txt"), "r").read()
email = Email("contato@mbkcontabilidade.com", PWD)

def get_client_data(cnpj):
    """Pega dados do cliente por meio das APIs em sequência."""
    endpoints = [
        f"http://192.168.0.105:8000/{cnpj}",
        f"http://192.168.0.114:8000/{cnpj}",
        f"https://minhareceita.org/{cnpj}"
    ]
    
    not_found_message = f"CNPJ {cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:14]} não encontrado."
    response = None
    
    # Tentar cada endpoint até encontrar um que retorne dados
    for endpoint in endpoints:
        try:
            response = request("GET", endpoint, timeout=2)
            response_data = json.loads(response.text)
            
            # Verificar se o CNPJ não foi encontrado
            if response_data.get("message") == not_found_message:
                logging.info(f"{not_found_message} no endpoint {endpoint}")
                continue  # Tentar o próximo endpoint
            
            logging.info(f"Dados obtidos da API para o CNPJ {cnpj}: {response_data}")
            break  # Se encontrou os dados, sair do loop
            
        except Exception as e:
            logging.error(f"Erro ao tentar obter dados para o CNPJ {cnpj} no endpoint {endpoint}: {e}")
            continue
    
    # Se nenhuma resposta válida foi obtida
    if not response or response_data.get("message") == not_found_message:
        logging.info(f"CNPJ {cnpj} não encontrado em nenhum endpoint. Pulando...")
        return None  # Retorna None para indicar que o CNPJ não foi encontrado
    
    # Processamento de dados se o CNPJ foi encontrado
    formatted_cnpj = f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:14]}"
    opcao_simples = response_data.get("opcao_pelo_simples", False)
    opcao_mei = response_data.get("opcao_pelo_mei", False)
    
    if opcao_simples and opcao_mei:
        frame = "MEI"
    elif opcao_simples:
        frame = "SIMPLES"
    elif opcao_mei:
        frame = "MEI"
    else:
        frame = " - "
    
    if "qsa" in response_data and response_data["qsa"]:
        responsable = response_data["qsa"][0]["nome_socio"]
    else:
        responsable = ""
    partners = [partner["nome_socio"] for partner in response_data.get("qsa", [])]
    
    partners = ", ".join(partners)
    year, month, day = response_data["data_inicio_atividade"].split("-")
    start_at = f"{day}/{month}/{year}"

    data = (
        [response_data["razao_social"], frame]
        + [None] * 2
        + [expires_at, partners, responsable]
        + [None]
        + [cnpj, formatted_cnpj]
        + [None] * 3
        + [response_data.get("ddd_telefone_1"), response_data.get("email")]
        + [None] * 13
        + [
            response_data.get("descricao_identificador_matriz_filial"),
            start_at,
            response_data.get("natureza_juridica"),
            response_data.get("uf"),
        ]
    )
    logging.info(f"Dados formatados para o CNPJ {cnpj}: {data}")
    return data

def process_ex_client(client):
    """Registrar ex-clientes"""
    old_client = False
    CLIENTES_INATIVOS_PF_PJ_XLSX = os.path.join(ROOT_DIR, "Clientes inativos - PJ e PF.xlsx")

    # Abrir a planilha de clientes inativos
    inativos_wb = load_workbook(CLIENTES_INATIVOS_PF_PJ_XLSX)
    inativos_ws = inativos_wb.active

    # Verificar se o cliente já é inativo
    for row in inativos_ws.iter_rows(values_only=True):
        if row[8] == client[8] or row[10] == client[10]:  # Verificar pelo CNPJ ou outro identificador
            logging.info(f"{row[0]} já foi nosso cliente")
            old_client = row
            break

    # Se não é inativo, adicionar o cliente à lista de inativos
    if not old_client:
        logging.info(f"O cliente {client[0]} foi movido para a lista de inativos")
        inativos_ws.append(client)

    # Salvar a planilha atualizada
    inativos_wb.save(CLIENTES_INATIVOS_PF_PJ_XLSX)


import os
import pandas as pd
import logging

def check_ex_client(client):
    """Verificar se o cliente supostamente novo já não foi um cliente antes"""
    old_client = False
    CLIENTES_INATIVOS_PF_PJ_XLSX = os.path.join(ROOT_DIR, "Clientes inativos - PJ e PF.xlsx")
    
    try:
        # Lendo o arquivo Excel
        df = pd.read_excel(CLIENTES_INATIVOS_PF_PJ_XLSX, engine='openpyxl')
        # Substituindo valores vazios por None
        df = df.where(pd.notnull(df), None)

        for _, ex_client in df.iterrows():
            if client == ex_client[8] or client == ex_client[10]:
                logging.info(f"{ex_client[0]} já foi nosso cliente")
                old_client = ex_client
                break

    except FileNotFoundError:
        logging.error(f"Arquivo {CLIENTES_INATIVOS_PF_PJ_XLSX} não encontrado.")
    except Exception as e:
        logging.error(f"Erro ao processar o arquivo: {e}")

    return old_client if old_client else False


def email_already_sent_today(log_filename, client_email):
    """Verifica se um email já foi enviado hoje para o cliente"""
    today_date = datetime.now().strftime('%Y-%m-%d')
    client_email = str(client_email)  # Garantir que client_email seja uma string
    logging.info(f"Verificando se e-mail já foi enviado hoje para {client_email}")

    if os.path.exists(log_filename):
        with open(log_filename, 'r') as log_file:
            for line in log_file:
                if today_date in line and client_email in line:
                    logging.info(f"E-mail já foi enviado hoje para {client_email}")
                    return True
    logging.info(f"Nenhum e-mail enviado hoje para {client_email}")
    return False

client_info_pj = {}
client_info_pf = {}
processed = set()
pfxs_directory = os.path.join(ROOT_DIR, "Certificados Digitais")
for subdir, dirs, files in os.walk(pfxs_directory):
    for file in files:
        if (
            file.endswith(".p12")
            and "VENCIDOS" not in subdir
            or file.endswith("pfx")
            and "VENCIDOS" not in subdir
        ):
            pfx_file = os.path.join(subdir, file)
            password = re.findall(r"\[(.*?)\]", file)[0]

            certificate = Certificate(pfx_file, password.encode())
            certificate.read_pfx_file()

            expires_at = certificate.not_valid_after()
            expires_at_str = expires_at.strftime("%d/%m/%Y")
            today = datetime.now()

            if expires_at.replace(tzinfo=None) < today:
                logging.info(f"Certificado vencido encontrado: {file}, movendo para pasta de vencidos")
                dest_path = os.path.join(VENCIDOS_DIR, file)
                shutil.move(pfx_file, dest_path)

                common_name = certificate.common_name()
                name, cnpj = common_name.split(":")

                if len(cnpj) > 11:
                    for row in cnpj_ws.iter_rows(min_row=2):
                        if row[8].value == cnpj:
                            inativos_ws.append([cell.value for cell in row])
                            cnpj_ws.delete_rows(row[0].row)
                            break

                else:
                    for row in cpf_ws.iter_rows(min_row=2):
                        if row[10].value == cnpj:
                            inativos_ws.append([cell.value for cell in row])
                            cpf_ws.delete_rows(row[0].row)
                            break

                continue
            
            expires_at = expires_at.strftime("%d/%m/%Y")
            commom_name = certificate.common_name()
            today = datetime.now()

            name, cnpj = commom_name.split(":")
            logging.info(f"Certificado encontrado: {name}, CNPJ/CPF: {cnpj}, expira em {expires_at}")

            if len(cnpj) > 11:
                if cnpj not in processed:
                    processed.add(cnpj)
                    if cnpj == "35419873000118":
                        continue
                    else:
                        api_data = get_client_data(cnpj)
                        client_info_pj[cnpj] = api_data
            else:
                if cnpj not in processed:
                    processed.add(cnpj)
                    formatted_cpf = f"{cnpj[:3]}.{cnpj[3:6]}.{cnpj[6:9]}-{cnpj[9:12]}"
                    client_info_pf[cnpj] = (
                        [name]
                        + [None] * 3
                        + [expires_at]
                        + [None] * 5
                        + [cnpj, formatted_cpf]
                        + [None] * 22
                    )

            for day_str, day in counter.items():
                if expires_at == day.strftime("%d/%m/%Y"):
                    logging.info(f"Certificado {name} vai expirar em {day_str} dias")
                    
                    phone_number = client_info_pj[cnpj][11]  # Assumindo que o número de telefone está no índice 11
                    # Remova quaisquer caracteres que não sejam dígitos
                    if phone_number:
                        phone_number = re.sub(r'\D', '', phone_number)
                        # Construa a URL com o número de telefone formatado
                        wa_link = f"https://wa.me/{phone_number}"
                    else:
                        wa_link = "https://wa.me/"  # Link vazio se não houver número de telefone
                        logging.warning(f"Número de telefone não disponível para o cliente {cnpj}")

                    # Cria a notificação
                    try:
                        toast = Notification(
                            app_id="MBK Contabilidade",
                            title=f"Certificado {name} vai expirar daqui a {day_str} dias!",
                            msg=f"O certificado expira no dia {expires_at}. Bora mandar mensagem avisando!",
                            duration="long",
                            icon=r"C:\Users\lucas\Desktop\MBK-app\mbkapp\static\main\logo_white.png",
                        )
                        toast.add_actions(label="Mandar mensagem", launch=wa_link)
                        toast.show()
                    except Exception as e:
                        logging.error(f"Falha ao exibir notificação: {e}")

                    client_email = process_file_info(client_info_pj[cnpj][0])
                    if not email_already_sent_today(log_filename, client_email):
                        email.send(
                            client_email,
                            f"O seu Certificado Digital vai expirar daqui a {day_str} dias!",
                            f"O certificado expira no dia {expires_at}. Vamos renovar!",
                        )
                        logging.info(f"E-mail enviado para {client_email} sobre expiração do certificado")
                    else:
                        logging.info(f"E-mail já enviado para {client_email} hoje. Não enviado novamente.")

found_in_inativos = False
for row in inativos_ws.iter_rows(min_row=2):
    if (len(cnpj) > 11 and row[8].value == cnpj) or (len(cnpj) <= 11 and row[10].value == cnpj):
        if len(cnpj) > 11:
            cnpj_ws.append([cell.value for cell in row])
        else:
            cpf_ws.append([cell.value for cell in row])

        inativos_ws.delete_rows(row[0].row)
        found_in_inativos = True
        logging.info(f"Cliente {name} ({cnpj}) movido de inativos para ativos.")
        break

# Criar listas de clientes do certificado e da tabela
pj_clients = sorted(list(client_info_pj.keys()))
pj_clients.append("35419873000118")  # É preciso que o CNPJ da MBK fique por último
client_info_pj["35419873000118"] = get_client_data("35419873000118")
pj_table_clients = sorted([str(cell.value) for cell in cnpj_ws["I"] if cell.value != None and cell.value != "CNPJ S/ PONTUAÇÃO"])

pf_clients = sorted(list(client_info_pf.keys()))
pf_table_clients = sorted([int(cell.value) for cell in cpf_ws["K"] if cell.value != None and cell.value != "CPF S/ PONTUAÇÃO"])

logging.info("Clientes PJ na tabela: " + ", ".join(pj_table_clients))
logging.info("Clientes PJ dos certificados: " + ", ".join(pj_clients))
logging.info("Clientes PF na tabela: " + ", ".join(map(str, pf_table_clients)))
logging.info("Clientes PF dos certificados: " + ", ".join(pf_clients))

# Associar todos os clientes às suas linhas
x = 0   
for cell in cnpj_ws["I"]:
    x += 1
    for cnpj in list(client_info_pj.keys()):
        if cnpj == cell.value:
            client_info_pj[cnpj] = [i.value for i in cnpj_ws[x]]
            for i in cnpj_ws[x]:
                i.value = None

x = 0
for cell in cpf_ws["K"]:
    x += 1
    for cpf in list(client_info_pf.keys()):
        if cpf == cell.value:
            client_info_pf[cpf] = [i.value for i in cpf_ws[x]]
            for i in cpf_ws[x]:
                i.value = None

collumns = list(string.ascii_uppercase)
collumns.extend(["AA", "AB", "AC", "AD", "AE", "AF"])

# Verificar diferença entre as duas listas criadas
if pj_clients != pj_table_clients:
    if len(pj_clients) > len(pj_table_clients):
        diff = list(set(pj_clients).difference(pj_table_clients))
        # verificar se cliente já foi cliente antes
        for client in diff:
            is_old_client = check_ex_client(client)
            if is_old_client:
                client_info_pj[client] = is_old_client
            else:
                continue
    elif len(pj_clients) < len(pj_table_clients):
        diff = list(set(pj_table_clients).difference(pj_clients))
        for client in diff:
            x = 0
            for cell in cnpj_ws["I"]:
                x += 1
                if client == cell.value:
                    ex_client = [i.value for i in cnpj_ws[x]]
                    process_ex_client(ex_client)
                    for i in cnpj_ws[x]:
                        i.value = None
                else:
                    continue
else:
    logging.info("Nenhuma diferença encontrada entre os clientes PJ dos certificados e da tabela")

if pf_clients != pf_table_clients:
    if len(pf_clients) > len(pf_table_clients):
        diff = list(set(pf_clients).difference(pf_table_clients))
        # verificar se cliente já foi cliente antes
        for client in diff:
            is_old_client = check_ex_client(client)
            if is_old_client:
                client_info_pf[client] = is_old_client
            else:
                continue
    elif len(pf_clients) < len(pf_table_clients):
        diff = list(set(pf_table_clients).difference(pf_clients))
        for client in diff:
            x = 0
            for cell in cpf_ws["K"]:
                x += 1
                if client == cell.value:
                    ex_client = [i.value for i in cpf_ws[x]]
                    process_ex_client(ex_client)
                    for i in cpf_ws[x]:
                        i.value = None
                else:
                    continue
else:
    logging.info("Nenhuma diferença encontrada entre os clientes PF dos certificados e da tabela")

i = 1
for client in list(client_info_pj.keys()):
    i += 1
    for collumn in collumns:
        cnpj_ws[f"{collumn}{i}"] = client_info_pj[client][collumns.index(collumn)]

i = 1
for client in list(client_info_pf.keys()):
    i += 1
    for collumn in collumns:
        cpf_ws[f"{collumn}{i}"] = client_info_pf[client][collumns.index(collumn)]

# Salvar os arquivos XLSX no diretório especificado
cpf_wb.save(os.path.join(XLSX_OUTPUT_DIR, "Clientes ativos - PF.xlsx"))
cnpj_wb.save(os.path.join(XLSX_OUTPUT_DIR, "Clientes ativos - PJ.xlsx"))
inativos_wb.save(os.path.join(XLSX_OUTPUT_DIR, "Clientes inativos - PJ e PF.xlsx"))
logging.info("Planilhas de clientes atualizadas com sucesso")
